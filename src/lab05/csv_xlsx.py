import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    p = Path(csv_path)
    if not p.exists():
        raise FileNotFoundError("файл не найден")
    if p.suffix.lower() != ".csv":
        raise ValueError("файл не является csv-файлом")
    if Path(xlsx_path).suffix.lower() != ".xlsx":
        raise ValueError("файл не является xlsx-файлом")
    with p.open("r", encoding="utf-8") as f:
        reader = csv.reader(f)
        data = list(reader)
    if not data:
        raise ValueError("файл пуст")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in data:
        sheet.append(row)
    for col_idx, col in enumerate(sheet.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = max((max_length + 2), 8)
        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    workbook.save(xlsx_path)

